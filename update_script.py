import openpyxl
import gspread
from oauth2client.service_account import ServiceAccountCredentials

######################################## Get data from the excel file from the local machine #########################################################

# load the excel file into a variable
wb = openpyxl.load_workbook('test.xlsx')

# get the specific sheet where your data resides
sheet = wb['Sheet1']

# instantiate 2 empty lists to get the data from two columns from excel
idList = []
statusList = []

# get data from first column and save it in the list
for i in range(1, sheet.max_row+1):
    idList.append(sheet.cell(row=i, column=1).value)

# get data from second column and save it in the other list
for i in range(1, sheet.max_row+1):
    statusList.append(sheet.cell(row=i, column=2).value)

####################################### Adding the data to Google Sheets #############################################################################

# Declare the scope = what will you use this for
scope = ['https://spreadsheets.google.com/feeds',
        "https://www.googleapis.com/auth/drive.file",
        'https://www.googleapis.com/auth/spreadsheets',
        "https://www.googleapis.com/auth/drive"]

# create credentials using the creds.json file downloaded from the google drive API and the scope
credentials = ServiceAccountCredentials.from_json_keyfile_name('creds.json', scope)

# Use gspread to authorise this script to access the sheets using the 'credentials' variable created above
client = gspread.authorize(credentials)


# Open the spreadsheet 
gSheet = client.open('test').sheet1


print(len(gSheet.col_values(1)))

# delete all values from the spreadsheet
for i in range(1, len(gSheet.col_values(1)) + 1):
    gSheet.update_cell(i, 1, '')
    gSheet.update_cell(i, 2, '')

# add the new data from the excel file saved in the lists
for i in range(0, len(idList)):
    gSheet.update_cell(i+1, 1, idList[i])
    gSheet.update_cell(i+1, 2, statusList[i])









